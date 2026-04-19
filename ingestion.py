from __future__ import annotations

from dataclasses import asdict, dataclass
from io import BytesIO
import hashlib
import json
import math
import re
from typing import Any

from openpyxl import load_workbook
import pandas as pd


KEYWORD_WEIGHTS = {
    "구분": 2.4,
    "전국": 1.5,
    "서울": 1.2,
    "매매": 1.0,
    "지수": 1.0,
}
HEADER_REGION_HINTS = ("전국", "서울", "경기", "인천", "부산", "대구", "광주", "대전", "울산", "세종")
DEFAULT_MIN_WEEKS = 104


@dataclass(frozen=True)
class QAReport:
    blocking_errors: tuple[str, ...]
    warnings: tuple[str, ...]
    stats: dict[str, Any]
    detected_layout: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def to_json(self) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=2, default=str)


def workbook_fingerprint(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def _coerce_text(value: Any) -> str:
    if _is_empty(value):
        return ""
    return str(value).strip()


def _parse_excel_date(value: Any) -> pd.Timestamp | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return pd.Timestamp(value).normalize()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value <= 0:
            return None
        try:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")).normalize()
        except (ValueError, OverflowError):
            return None
    text = _coerce_text(value)
    if not text:
        return None
    text = text.replace("/", ".").replace("-", ".")
    text = re.sub(r"\s+", "", text)
    for candidate in (text, text.replace(".", "-"), text.replace(".", "")):
        parsed = pd.to_datetime(candidate, errors="coerce")
        if not pd.isna(parsed):
            return pd.Timestamp(parsed).normalize()
    return None


def _monday_floor(timestamp: pd.Timestamp) -> pd.Timestamp:
    return (timestamp - pd.Timedelta(days=int(timestamp.weekday()))).normalize()


def _series_date_ratio(values: list[Any]) -> float:
    non_empty = [value for value in values if not _is_empty(value)]
    if not non_empty:
        return 0.0
    parsed = sum(_parse_excel_date(value) is not None for value in non_empty)
    return parsed / len(non_empty)


def _series_numeric_ratio(values: list[Any]) -> float:
    candidates = [value for value in values if not _is_empty(value)]
    if not candidates:
        return 0.0
    numeric = 0
    for value in candidates:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numeric += 1
            continue
        text = _coerce_text(value).replace(",", "")
        converted = pd.to_numeric(text, errors="coerce")
        numeric += int(pd.notna(converted))
    return numeric / len(candidates)


def detect_kb_sheet(workbook: Any) -> dict[str, Any]:
    sheet_scores: list[dict[str, Any]] = []
    best: dict[str, Any] | None = None
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=15, values_only=True):
            rows.append(list(row))
        flattened = [_coerce_text(value) for row in rows for value in row if not _is_empty(value)]
        joined = " ".join(flattened)
        keyword_score = sum(weight for keyword, weight in KEYWORD_WEIGHTS.items() if keyword in joined)
        first_col = [row[0] for row in rows if row]
        rest = [value for row in rows for value in row[1:]]
        date_ratio = _series_date_ratio(first_col)
        numeric_density = _series_numeric_ratio(rest)
        visible_bonus = 0.35 if getattr(sheet, "sheet_state", "visible") == "visible" else -0.2
        header_hint = 0.25 if "구분" in joined or sum(hint in joined for hint in HEADER_REGION_HINTS) >= 2 else 0.0
        score = keyword_score + (date_ratio * 2.0) + numeric_density + visible_bonus + header_hint
        result = {
            "sheet_name": sheet.title,
            "sheet_state": getattr(sheet, "sheet_state", "visible"),
            "score": round(score, 4),
            "keyword_score": round(keyword_score, 4),
            "date_ratio": round(date_ratio, 4),
            "numeric_density": round(numeric_density, 4),
        }
        sheet_scores.append(result)
        if best is None or result["score"] > best["score"]:
            best = result
    return {"best_sheet": best, "sheet_scores": sheet_scores}


def _make_unique_headers(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for idx, value in enumerate(values):
        base = _coerce_text(value) or f"column_{idx}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers


def _score_header_row(row: pd.Series) -> float:
    texts = [_coerce_text(value) for value in row.tolist()]
    joined = " ".join(texts)
    score = 0.0
    if "구분" in texts:
        score += 3.0
    score += 0.8 * sum(hint in joined for hint in HEADER_REGION_HINTS)
    string_count = sum(bool(text) for text in texts)
    date_like = sum(_parse_excel_date(value) is not None for value in row.tolist())
    numeric_like = sum(pd.notna(pd.to_numeric(_coerce_text(value).replace(",", ""), errors="coerce")) for value in row.tolist())
    score += min(string_count, 12) * 0.08
    score -= date_like * 0.2
    score -= numeric_like * 0.05
    return score


def _detect_header_row(raw_df: pd.DataFrame) -> int:
    sample = raw_df.head(8)
    scores = {int(index): _score_header_row(sample.loc[index]) for index in sample.index}
    return max(scores, key=scores.get)


def normalize_kb_sheet(
    raw_df: pd.DataFrame,
    *,
    source_file: str,
    sheet_name: str,
    min_weeks: int = DEFAULT_MIN_WEEKS,
) -> tuple[pd.DataFrame, QAReport]:
    blocking_errors: list[str] = []
    warnings: list[str] = []
    stats: dict[str, Any] = {}
    working = raw_df.dropna(axis=0, how="all").dropna(axis=1, how="all").reset_index(drop=True)
    if working.empty:
        report = QAReport(("The detected worksheet is empty.",), tuple(), {}, {"sheet_name": sheet_name})
        return pd.DataFrame(columns=["date", "region", "value", "is_imputed", "source_file", "sheet_name"]), report

    header_row = _detect_header_row(working)
    headers = _make_unique_headers(working.iloc[header_row].tolist())
    data = working.iloc[header_row + 1 :].copy()
    data.columns = headers
    data = data.dropna(axis=0, how="all").reset_index(drop=True)
    if data.empty:
        report = QAReport(("The detected worksheet has no rows below the header.",), tuple(), {}, {"sheet_name": sheet_name})
        return pd.DataFrame(columns=["date", "region", "value", "is_imputed", "source_file", "sheet_name"]), report

    date_col = "구분" if "구분" in data.columns else data.columns[0]
    parsed_dates = data[date_col].map(_parse_excel_date)
    valid_row_mask = parsed_dates.notna()
    if not valid_row_mask.any():
        blocking_errors.append("No valid date column could be parsed from the worksheet.")
    data = data.loc[valid_row_mask].copy()
    data["date"] = parsed_dates.loc[valid_row_mask].map(_monday_floor)
    data = data.drop(columns=[date_col], errors="ignore")

    candidate_columns = [column for column in data.columns if column != "date" and not str(column).startswith("Unnamed")]
    if not candidate_columns:
        blocking_errors.append("No usable region columns were detected.")

    melted_frames: list[pd.DataFrame] = []
    numeric_failure_count = 0
    numeric_input_count = 0
    duplicate_conflicts = 0
    imputed_points = 0
    max_gap = 0

    for column in candidate_columns:
        region_frame = pd.DataFrame({"date": data["date"], "raw_value": data[column]})
        region_frame = region_frame.loc[~region_frame["raw_value"].map(_is_empty)].copy()
        if region_frame.empty:
            continue
        cleaned = region_frame["raw_value"].map(lambda value: _coerce_text(value).replace(",", ""))
        region_frame["value"] = pd.to_numeric(cleaned, errors="coerce")
        numeric_input_count += int(region_frame["raw_value"].notna().sum())
        numeric_failure_count += int(region_frame["value"].isna().sum())
        dedup = region_frame.groupby("date")["value"].agg(list).reset_index()
        dedup["unique_non_null"] = dedup["value"].map(lambda values: sorted({round(v, 12) for v in values if pd.notna(v)}))
        conflicts = dedup["unique_non_null"].map(len).gt(1).sum()
        duplicate_conflicts += int(conflicts)
        if conflicts:
            blocking_errors.append(f"Region '{column}' contains conflicting duplicate weekly values.")
            continue
        dedup["value"] = dedup["unique_non_null"].map(lambda values: values[0] if values else pd.NA)
        dedup = dedup.loc[dedup["value"].notna(), ["date", "value"]].sort_values("date")
        if dedup.empty:
            continue
        full_index = pd.date_range(dedup["date"].min(), dedup["date"].max(), freq="W-MON")
        expanded = dedup.set_index("date").reindex(full_index)
        expanded.index.name = "date"
        missing_mask = expanded["value"].isna()
        if missing_mask.any():
            run_ids = missing_mask.ne(missing_mask.shift(fill_value=False)).cumsum()
            run_lengths = missing_mask.groupby(run_ids).sum()
            true_run_lengths = [int(length) for length, is_missing in zip(run_lengths, missing_mask.groupby(run_ids).first()) if is_missing]
            max_gap = max(max_gap, max(true_run_lengths, default=0))
            if any(length >= 3 for length in true_run_lengths):
                blocking_errors.append(
                    f"Region '{column}' contains a missing weekly run of 3 weeks or longer after reindexing."
                )
            elif true_run_lengths:
                warnings.append(f"Region '{column}' contains short gaps that were linearly interpolated.")
                expanded["value"] = expanded["value"].interpolate(method="linear", limit=2, limit_direction="both")
        expanded["is_imputed"] = missing_mask & expanded["value"].notna()
        imputed_points += int(expanded["is_imputed"].sum())
        if expanded["value"].dropna().shape[0] < min_weeks:
            blocking_errors.append(
                f"Region '{column}' has only {expanded['value'].dropna().shape[0]} weekly observations; {min_weeks} are required."
            )
        expanded = expanded.reset_index().rename(columns={"index": "date"})
        expanded["region"] = column
        expanded["source_file"] = source_file
        expanded["sheet_name"] = sheet_name
        melted_frames.append(expanded[["date", "region", "value", "is_imputed", "source_file", "sheet_name"]])

    if numeric_input_count:
        failure_ratio = numeric_failure_count / numeric_input_count
        stats["numeric_failure_ratio"] = round(failure_ratio, 4)
        if failure_ratio > 0.2:
            blocking_errors.append("Numeric conversion failure ratio exceeded 20% for the uploaded worksheet.")
        elif failure_ratio > 0:
            warnings.append("Some worksheet cells could not be converted into numeric values.")
    else:
        stats["numeric_failure_ratio"] = 0.0

    if duplicate_conflicts:
        stats["duplicate_conflicts"] = duplicate_conflicts
    if not melted_frames:
        blocking_errors.append("No valid region series could be normalized from the worksheet.")
        normalized = pd.DataFrame(columns=["date", "region", "value", "is_imputed", "source_file", "sheet_name"])
    else:
        normalized = pd.concat(melted_frames, ignore_index=True).sort_values(["region", "date"]).reset_index(drop=True)

    unique_weeks = normalized["date"].nunique() if not normalized.empty else 0
    if unique_weeks and unique_weeks < min_weeks:
        blocking_errors.append(f"The workbook contains only {unique_weeks} unique weekly observations; {min_weeks} are required.")

    stats.update(
        {
            "rows": int(normalized.shape[0]),
            "regions": int(normalized["region"].nunique()) if not normalized.empty else 0,
            "unique_weeks": int(unique_weeks),
            "imputed_points": int(imputed_points),
            "max_missing_run": int(max_gap),
            "header_row": int(header_row),
            "date_column": date_col,
        }
    )
    detected_layout = {
        "sheet_name": sheet_name,
        "header_row": int(header_row),
        "date_column": date_col,
        "columns": headers,
    }
    report = QAReport(tuple(dict.fromkeys(blocking_errors)), tuple(dict.fromkeys(warnings)), stats, detected_layout)
    return normalized, report


def load_kb_workbook(file_bytes: bytes, source_name: str) -> tuple[pd.DataFrame, QAReport]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    detection = detect_kb_sheet(workbook)
    best_sheet = detection.get("best_sheet")
    if best_sheet is None:
        report = QAReport(("No worksheet matched KB weekly price-index heuristics.",), tuple(), {}, detection)
        return pd.DataFrame(columns=["date", "region", "value", "is_imputed", "source_file", "sheet_name"]), report
    raw_df = pd.read_excel(BytesIO(file_bytes), sheet_name=best_sheet["sheet_name"], header=None, engine="openpyxl")
    normalized, qa_report = normalize_kb_sheet(raw_df, source_file=source_name, sheet_name=best_sheet["sheet_name"])
    combined_layout = dict(qa_report.detected_layout)
    combined_layout["sheet_scores"] = detection["sheet_scores"]
    combined_layout["fingerprint"] = workbook_fingerprint(file_bytes)
    return normalized, QAReport(
        qa_report.blocking_errors,
        qa_report.warnings,
        qa_report.stats,
        combined_layout,
    )
